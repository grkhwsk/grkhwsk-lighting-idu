"""
passport_s0.py
Паспорт световой среды — характеристики маски S0 по кварталам.

Запуск:
    python scripts/passport_s0.py --city norilsk
    python scripts/passport_s0.py --city polyarnye_zori

Выход: output/passport_s0.xlsx  (три листа: сводная, кварталы, POI по зонам)
"""

import geopandas as gpd
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils import load_config, get_path, get_output_dir

# ── Конфиг и пути ────────────────────────────────────────────────

config, BASE = load_config()

BLOCKS_FILE = get_path(config, BASE, "blocks_raw")
OUTPUT_FILE = get_output_dir(config, BASE) / "passport_s0.xlsx"

# ── Константы модели ─────────────────────────────────────────────
# Типы активности и порядок зон одинаковы для всех городов

TYPE_LABELS = {
    "child_activity":          "Детская активность",
    "education":               "Образование",
    "sport":                   "Спорт",
    "recreation":              "Рекреация",
    "services_retail":         "Услуги и торговля",
    "administration_business": "Управление",
    "healthcare":              "Здравоохранение",
    "leisure_culture":         "Культура и досуг",
    "transit":                 "Транспортные узлы",
}
TYPE_COLS  = list(TYPE_LABELS.keys())
ZONE_ORDER = ["Public", "Residential", "Special_Restricted",
              "Non_Urban", "Industrial", "Outside"]

# ── Стили ────────────────────────────────────────────────────────

H_BG, H_FG = "2E4057", "FFFFFF"
Z_BG, Z_FG = "D6E4F0", "1A2E44"
_side = Side(style="thin", color="CCCCCC")


def hfill(hex_):   return PatternFill("solid", fgColor=hex_)
def fnt(bold=False, color="000000", sz=10):
    return Font(name="Arial", bold=bold, color=color, size=sz)
def aln(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def brd():
    return Border(left=_side, right=_side, top=_side, bottom=_side)


def mask_fill(val):
    """Цветовая шкала по % покрытия маской."""
    if not isinstance(val, (int, float)) or pd.isna(val): return hfill("FFFFFF")
    if val >= 50: return hfill("FFD6D6")
    if val >= 25: return hfill("FFF3CD")
    if val > 0:   return hfill("F5F5F5")
    return hfill("FFFFFF")


def p24h_fill(val):
    """Цветовая шкала по % постоянной потребности."""
    if not isinstance(val, (int, float)) or pd.isna(val): return hfill("FFFFFF")
    if val >= 50: return hfill("FFD6D6")
    if val >= 20: return hfill("FFF3CD")
    if val > 0:   return hfill("F5F5F5")
    return hfill("FFFFFF")


def header_row(ws, row_num, values, widths):
    for ci, (val, w) in enumerate(zip(values, widths), 1):
        c = ws.cell(row=row_num, column=ci, value=val)
        c.font      = fnt(bold=True, color=H_FG)
        c.fill      = hfill(H_BG)
        c.alignment = aln("center", wrap=True)
        c.border    = brd()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row_num].height = 32


def zone_header(ws, row_num, label, ncols):
    for ci in range(1, ncols + 1):
        c = ws.cell(row=row_num, column=ci)
        c.fill = hfill(Z_BG)
        c.font = fnt(bold=True, color=Z_FG)
        c.border = brd()
    ws.cell(row=row_num, column=1).value = f"  {label}"
    ws.row_dimensions[row_num].height = 16


def write_cell(ws, row_num, col, val, bold=False, center=True,
               fill=None, num_fmt=None):
    c = ws.cell(row=row_num, column=col, value=val)
    c.font      = fnt(bold=bold)
    c.alignment = aln("center" if center else "left")
    c.border    = brd()
    if fill:    c.fill = fill
    if num_fmt: c.number_format = num_fmt
    return c

# ── Загрузка данных ───────────────────────────────────────────────

print(f"\nЧитаю {BLOCKS_FILE}...")

if not BLOCKS_FILE.exists():
    raise FileNotFoundError(
        f"Файл кварталов не найден: {BLOCKS_FILE}\n"
        f"Сначала запустите blocks.py --city {BASE.name}"
    )

gdf = gpd.read_file(BLOCKS_FILE)

# Площадь в гектарах
gdf["area_ha"] = (
    gdf["area_m2"] / 10000 if "area_m2" in gdf.columns
    else gdf.geometry.area / 10000
).round(1)

# Заполнение пропусков
gdf["zone_group"]    = gdf["zone_group"].fillna("Outside")
gdf["zone_code"]     = gdf["zone_code"].fillna("—")
gdf["poi_count"]     = gdf["poi_count"].fillna(0).astype(int)
gdf["dominant_type"] = gdf["dominant_type"].fillna("—")
gdf["mask_pct"]      = gdf["mask_pct"].fillna(0)

# pct_24h: если колонка есть — заполняем NaN, если нет — создаём нулевую
if "pct_24h" in gdf.columns:
    gdf["pct_24h"] = gdf["pct_24h"].fillna(0)
else:
    gdf["pct_24h"] = 0.0

# POI-поля и порядок зон
poi_type_fields = [f"poi_{t}" for t in TYPE_COLS if f"poi_{t}" in gdf.columns]
zones_present   = [z for z in ZONE_ORDER if z in gdf["zone_group"].values]
zones_present  += [z for z in gdf["zone_group"].unique() if z not in ZONE_ORDER]

print(f"  Кварталов: {len(gdf)}, зон: {len(zones_present)}")

# ── Workbook ──────────────────────────────────────────────────────

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════
# ЛИСТ 1 — Сводная по зонам
# ════════════════════════════════════════════════════

ws1 = wb.active
ws1.title = "Сводная по зонам"

h1 = ["Зона ПЗЗ", "Кварталов", "Площадь, км²",
      "POI rank 1", "% покрытия маской", "Пост. потребность, %"]
header_row(ws1, 1, h1, [24, 12, 14, 12, 20, 22])

ri = 2
for zone in zones_present:
    zdf = gdf[gdf["zone_group"] == zone]
    if zdf.empty: continue

    row_vals = [
        zone, len(zdf),
        round(zdf["area_ha"].sum() / 100, 2),
        int(zdf["poi_count"].sum()),
        round(zdf["mask_pct"].mean(), 1),
        round(zdf["pct_24h"].mean(), 1),
    ]
    for ci, val in enumerate(row_vals, 1):
        fill = nfmt = None
        if ci == 3: nfmt = "0.00"
        if ci == 5: fill = mask_fill(val); nfmt = "0.0"
        if ci == 6: fill = p24h_fill(val); nfmt = "0.0"
        write_cell(ws1, ri, ci, val, bold=(ci == 1), center=(ci != 1),
                   fill=fill, num_fmt=nfmt)
    ri += 1

ws1.freeze_panes = "A2"
print('  Лист "Сводная по зонам" — готов')

# ════════════════════════════════════════════════════
# ЛИСТ 2 — Кварталы детально
# ════════════════════════════════════════════════════

ws2 = wb.create_sheet("Кварталы детально")

poi_type_labels = [
    TYPE_LABELS.get(f.replace("poi_", ""), f.replace("poi_", ""))
    for f in poi_type_fields
]
h2 = (["Зона ПЗЗ", "№ квартала", "Код зоны", "Площадь, га",
        "POI rank 1", "Тип активности", "% покрытия маской",
        "Пост. потребность, %"] + poi_type_labels)
header_row(ws2, 1, h2, [22, 12, 10, 12, 12, 26, 20, 22] + [13] * len(poi_type_fields))

ri, prev_zone = 2, None
for zone in zones_present:
    zdf = gdf[gdf["zone_group"] == zone].sort_values("poi_count", ascending=False)
    if zdf.empty: continue

    if zone != prev_zone:
        zone_header(ws2, ri,
                    f"ЗОНА: {zone}  ({len(zdf)} кв.  |  POI rank 1: {int(zdf['poi_count'].sum())}  |"
                    f"  Маска ср.: {round(zdf['mask_pct'].mean(), 1)}%",
                    len(h2))
        ri += 1
        prev_zone = zone

    for _, row in zdf.iterrows():
        dom       = TYPE_LABELS.get(row["dominant_type"], row["dominant_type"])
        mask      = round(float(row["mask_pct"]), 1)
        p24h      = round(float(row["pct_24h"]), 1)
        type_vals = [int(row[f]) if f in row and pd.notna(row[f]) else 0
                     for f in poi_type_fields]

        row_vals = ([zone, int(row["block_id"]), row["zone_code"],
                     row["area_ha"], int(row["poi_count"]), dom,
                     mask, p24h] + type_vals)

        for ci, val in enumerate(row_vals, 1):
            fill = nfmt = None
            center = ci not in (1, 6)
            if ci == 4: nfmt = "0.0"
            if ci == 7: fill = mask_fill(val); nfmt = "0.0"
            if ci == 8: fill = p24h_fill(val); nfmt = "0.0"
            write_cell(ws2, ri, ci, val, center=center, fill=fill, num_fmt=nfmt)
        ri += 1

ws2.freeze_panes = "A2"
print('  Лист "Кварталы детально" — готов')

# ════════════════════════════════════════════════════
# ЛИСТ 3 — Состав POI по зонам
# ════════════════════════════════════════════════════

ws3 = wb.create_sheet("Состав POI по зонам")
header_row(ws3, 1,
           ["Зона ПЗЗ", "Тип активности", "Объектов rank 1", "Доля в зоне, %"],
           [24, 30, 18, 16])

ri = 2
for zone in zones_present:
    zdf = gdf[gdf["zone_group"] == zone]
    if zdf.empty: continue

    type_totals = {
        f.replace("poi_", ""): int(zdf[f].fillna(0).sum())
        for f in poi_type_fields
    }
    type_totals = {k: v for k, v in type_totals.items() if v > 0}
    if not type_totals: continue

    total = sum(type_totals.values())
    zone_header(ws3, ri, zone, 4)
    ri += 1

    for t, n in sorted(type_totals.items(), key=lambda x: -x[1]):
        pct      = round(n / total * 100, 1)
        row_vals = [zone, TYPE_LABELS.get(t, t), n, pct]
        for ci, val in enumerate(row_vals, 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.font      = fnt(sz=10, color="999999" if ci == 1 else "000000")
            c.border    = brd()
            c.alignment = aln("center" if ci in (3, 4) else "left")
            if ci == 4: c.number_format = "0.0"
        ri += 1

ws3.freeze_panes = "A2"
print('  Лист "Состав POI по зонам" — готов')

# ── Сохранение ────────────────────────────────────────────────────

wb.save(OUTPUT_FILE)
print(f"\nГотово: {OUTPUT_FILE}")