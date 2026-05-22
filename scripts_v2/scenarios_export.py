"""
scenarios_export.py
Экспорт результатов сценарного расчёта NeedIndex в xlsx.

Запуск:
    python scripts/scenarios_export.py --city norilsk
    python scripts/scenarios_export.py --city polyarnye_zori

Изменения относительно оригинала:
  - добавлен аргумент --city
  - пути к файлам берутся из config.yaml
  - K_time берётся из output/k_time_table.csv (не захардкожен)
  - K_albedo и пороги берутся из config.yaml
  - в шапку параметров добавлена строка "Город"
"""

import geopandas as gpd
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils import load_config, get_path, get_output_dir

# ── Конфиг и пути ────────────────────────────────────────────────

config, BASE = load_config()

BLOCKS_FILE = get_path(config, BASE, "blocks")
OUTPUT_FILE = get_output_dir(config, BASE) / "scenarios.xlsx"
K_TIME_FILE = get_output_dir(config, BASE) / "k_time_table.csv"

# ── K_time из CSV, K_albedo из конфига ───────────────────────────

k_time_df  = pd.read_csv(K_TIME_FILE)
albedo_cfg = config["albedo"]["scenarios"]
tau1       = config["thresholds"]["tau_1"]
tau2       = config["thresholds"]["tau_2"]


def get_k_time(season, slot, scenario="scenario_1"):
    row = k_time_df[
        (k_time_df["season"]        == season) &
        (k_time_df["slot"]          == slot)   &
        (k_time_df["time_scenario"] == scenario)
    ]
    return float(row["K_time"].values[0]) if len(row) > 0 else 0.0


def get_k_albedo(albedo):
    return albedo_cfg[albedo]["K_albedo"]


# ── Список сценариев ─────────────────────────────────────────────

SCENARIOS = [
    # S0 — базовый
    {"label":   "S0 — Базовый (межсезонье / вечер / no_snow)",
     "sheet":   "S0 базовый",
     "season":  "midseason", "slot": "evening", "albedo": "no_snow",
     "k_time":  get_k_time("midseason", "evening"),
     "k_albedo": get_k_albedo("no_snow"),
     "ni_field": "ni_S0_midseason_evening", "pct_field": "pct_S0_midseason_evening"},

    # S1 — варьируется слот
    {"label":   "S1 — Межсезонье / Утро / no_snow",
     "sheet":   "S1 утро",
     "season":  "midseason", "slot": "morning", "albedo": "no_snow",
     "k_time":  get_k_time("midseason", "morning"),
     "k_albedo": get_k_albedo("no_snow"),
     "ni_field": "ni_S1_midseason_morning", "pct_field": "pct_S1_midseason_morning"},

    {"label":   "S1 — Межсезонье / День / no_snow",
     "sheet":   "S1 день",
     "season":  "midseason", "slot": "day", "albedo": "no_snow",
     "k_time":  get_k_time("midseason", "day"),
     "k_albedo": get_k_albedo("no_snow"),
     "ni_field": "ni_S1_midseason_day", "pct_field": "pct_S1_midseason_day"},

    {"label":   "S1 — Межсезонье / Ночь / no_snow",
     "sheet":   "S1 ночь",
     "season":  "midseason", "slot": "night", "albedo": "no_snow",
     "k_time":  get_k_time("midseason", "night"),
     "k_albedo": get_k_albedo("no_snow"),
     "ni_field": "ni_S1_midseason_night", "pct_field": "pct_S1_midseason_night"},

    # S2 — варьируется сезон
    {"label":   "S2 — Зима / Вечер / no_snow",
     "sheet":   "S2 зима",
     "season":  "winter", "slot": "evening", "albedo": "no_snow",
     "k_time":  get_k_time("winter", "evening"),
     "k_albedo": get_k_albedo("no_snow"),
     "ni_field": "ni_S2_winter_evening", "pct_field": "pct_S2_winter_evening"},

    {"label":   "S2 — Лето / Вечер / no_snow",
     "sheet":   "S2 лето",
     "season":  "summer", "slot": "evening", "albedo": "no_snow",
     "k_time":  get_k_time("summer", "evening"),
     "k_albedo": get_k_albedo("no_snow"),
     "ni_field": "ni_S2_summer_evening", "pct_field": "pct_S2_summer_evening"},

    # S3 — варьируется альбедо
    {"label":   "S3 — Межсезонье / Вечер / dirty_snow",
     "sheet":   "S3 грязный снег",
     "season":  "midseason", "slot": "evening", "albedo": "dirty_snow",
     "k_time":  get_k_time("midseason", "evening"),
     "k_albedo": get_k_albedo("dirty_snow"),
     "ni_field": "ni_S3_midseason_evening_dirty", "pct_field": "pct_S3_midseason_evening_dirty"},

    {"label":   "S3 — Межсезонье / Вечер / clean_snow",
     "sheet":   "S3 чистый снег",
     "season":  "midseason", "slot": "evening", "albedo": "clean_snow",
     "k_time":  get_k_time("midseason", "evening"),
     "k_albedo": get_k_albedo("clean_snow"),
     "ni_field": "ni_S3_midseason_evening_clean", "pct_field": "pct_S3_midseason_evening_clean"},

    # Сценарии весов K_time
    {"label":   "S0 — Межсезонье / Вечер / scenario_2",
     "sheet":   "S0 сценарий 2",
     "season":  "midseason", "slot": "evening", "albedo": "no_snow",
     "k_time":  get_k_time("midseason", "evening", "scenario_2"),
     "k_albedo": get_k_albedo("no_snow"),
     "ni_field": "ni_S0_midseason_evening_sc2", "pct_field": "pct_S0_midseason_evening_sc2"},

    {"label":   "S0 — Межсезонье / Вечер / scenario_3",
     "sheet":   "S0 сценарий 3",
     "season":  "midseason", "slot": "evening", "albedo": "no_snow",
     "k_time":  get_k_time("midseason", "evening", "scenario_3"),
     "k_albedo": get_k_albedo("no_snow"),
     "ni_field": "ni_S0_midseason_evening_sc3", "pct_field": "pct_S0_midseason_evening_sc3"},
]

NI_CLASSES = {
    0: "нет потребности",
    1: f"низкая (0–{tau1})",
    2: f"средняя ({tau1}–{tau2})",
    3: f"высокая (>{tau2})",
}

ACTIVE_TYPES = {
    "morning": ["child_activity", "education", "sport", "recreation",
                "services_retail", "administration_business",
                "healthcare", "transit", "leisure_culture"],
    "day":     ["child_activity", "education", "sport", "recreation",
                "services_retail", "administration_business",
                "healthcare", "transit", "leisure_culture"],
    "evening": ["child_activity", "sport", "recreation",
                "services_retail", "administration_business",
                "healthcare", "transit", "leisure_culture"],
    "night":   ["administration_business", "healthcare", "transit"],
}

ZONE_ORDER = ["Public", "Residential", "Special_Restricted",
              "Non_Urban", "Industrial", "Outside"]

# ── Стили ────────────────────────────────────────────────────────

H_BG, H_FG = "2E4057", "FFFFFF"
P_BG        = "F0F4F8"
Z_BG, Z_FG = "D6E4F0", "1A2E44"
_side = Side(style="thin", color="CCCCCC")


def hfill(h):  return PatternFill("solid", fgColor=h)
def fnt(bold=False, color="000000", sz=10):
    return Font(name="Arial", bold=bold, color=color, size=sz)
def aln(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def brd():
    return Border(left=_side, right=_side, top=_side, bottom=_side)


def ni_fill(val):
    if not isinstance(val, (int, float)) or pd.isna(val): return hfill("FFFFFF")
    if val >= tau2:  return hfill("FFD6D6")
    if val >= tau1:  return hfill("FFF3CD")
    if val > 0:      return hfill("F5F5F5")
    return hfill("FFFFFF")


def header_row(ws, row_num, values, widths):
    for ci, (val, w) in enumerate(zip(values, widths), 1):
        c = ws.cell(row=row_num, column=ci, value=val)
        c.font = fnt(bold=True, color=H_FG)
        c.fill = hfill(H_BG)
        c.alignment = aln("center", wrap=True)
        c.border = brd()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row_num].height = 28


def zone_header(ws, row_num, label, ncols):
    for ci in range(1, ncols + 1):
        c = ws.cell(row=row_num, column=ci)
        c.fill = hfill(Z_BG); c.font = fnt(bold=True, color=Z_FG); c.border = brd()
    ws.cell(row=row_num, column=1).value = f"  {label}"
    ws.row_dimensions[row_num].height = 16


def param_row(ws, row_num, label, value, ncols=6):
    for ci, val in enumerate([label, value] + [""] * (ncols - 2), 1):
        c = ws.cell(row=row_num, column=ci, value=val)
        c.fill = hfill(P_BG); c.border = brd()
        c.font = fnt(bold=(ci == 1)); c.alignment = aln("left")


def active_poi_count(row, slot, poi_type_cols):
    types = ACTIVE_TYPES.get(slot, [])
    return sum(
        int(row[col]) for col in poi_type_cols
        if col.replace("poi_", "") in types and col in row and pd.notna(row[col])
    )


# ── Данные ────────────────────────────────────────────────────────

print(f"\nЧитаю {BLOCKS_FILE}...")
gdf = gpd.read_file(BLOCKS_FILE)

gdf["area_ha"]       = (gdf.geometry.area / 10000).round(1)
gdf["zone_group"]    = gdf["zone_group"].fillna("Outside")
gdf["poi_count"]     = gdf["poi_count"].fillna(0).astype(int)
gdf["dominant_type"] = gdf["dominant_type"].fillna("—")
poi_type_cols  = [c for c in gdf.columns if c.startswith("poi_") and c != "poi_count"]
zones_present  = [z for z in ZONE_ORDER if z in gdf["zone_group"].values]
zones_present += [z for z in gdf["zone_group"].unique() if z not in ZONE_ORDER]

# ── Workbook ──────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)

for sc in SCENARIOS:
    ws    = wb.create_sheet(sc["sheet"])
    ni_f  = sc["ni_field"]
    pct_f = sc["pct_field"]

    params = [
        ("Сценарий",  sc["label"]),
        ("Город",     config["location"]["name"]),
        ("Сезон",     sc["season"]),
        ("Слот",      sc["slot"]),
        ("K_time",    sc["k_time"]),
        ("K_albedo",  f"{sc['k_albedo']} ({sc['albedo']})"),
    ]
    for i, (lbl, val) in enumerate(params, 1):
        param_row(ws, i, lbl, val)
        ws.row_dimensions[i].height = 16
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30

    h = ["Зона ПЗЗ", "№ квартала", "Площадь, га", "POI rank 1 всего",
         "POI активных в слоте", "Доминирующий тип",
         "NeedIndex (ср.)", "% активных ячеек", "Класс потребности"]
    header_row(ws, 7, h, [22, 12, 12, 14, 20, 26, 16, 18, 24])

    ri = 8
    for zone in zones_present:
        zdf = gdf[gdf["zone_group"] == zone].copy()
        if zdf.empty: continue
        ni_mean = zdf[ni_f].mean() if ni_f in gdf.columns else 0
        zone_header(ws, ri,
                    f"ЗОНА: {zone}  ({len(zdf)} кв.)  |  NI ср.: {round(ni_mean, 3)}",
                    len(h))
        ri += 1
        for _, row in zdf.sort_values("block_id").iterrows():
            ni_val  = round(float(row[ni_f]),  3) if ni_f  in row and pd.notna(row[ni_f])  else 0.0
            pct_val = round(float(row[pct_f]), 1) if pct_f in row and pd.notna(row[pct_f]) else 0.0
            if ni_val == 0:         ni_class = NI_CLASSES[0]
            elif ni_val <= tau1:    ni_class = NI_CLASSES[1]
            elif ni_val <= tau2:    ni_class = NI_CLASSES[2]
            else:                   ni_class = NI_CLASSES[3]

            row_vals = [zone, int(row["block_id"]), row["area_ha"],
                        int(row["poi_count"]),
                        active_poi_count(row, sc["slot"], poi_type_cols),
                        row["dominant_type"], ni_val, pct_val, ni_class]

            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = fnt(sz=10); c.border = brd()
                c.alignment = aln("center" if ci not in (1, 6, 9) else "left")
                if ci == 3: c.number_format = "0.0"
                if ci == 7: c.fill = ni_fill(val); c.number_format = "0.000"
                if ci == 8: c.number_format = "0.0"
            ri += 1

    ws.freeze_panes = "A8"
    print(f'  Лист "{sc["sheet"]}" — готов')

# Лист легенды
wl = wb.create_sheet("Легенда")
wl.column_dimensions["A"].width = 28
wl.column_dimensions["B"].width = 50

legend_rows = [
    ("ПАРАМЕТРЫ РАСЧЁТА", ""),
    ("NeedIndex", "Индекс потребности в освещении: NI = Mask × Active × K_time × K_albedo"),
    ("Mask",      "Принадлежность ячейки к зоне потребности (0 или 1)"),
    ("Active",    "Активность объектов в данный временной слот (0 или 1)"),
    ("K_time",    "Доля тёмного времени в слоте для данного сезона"),
    ("K_albedo",  "Коэффициент альбедо поверхности"),
    ("", ""),
    ("КЛАССЫ ПОТРЕБНОСТИ", ""),
    ("Класс 0", "NI = 0 — потребность отсутствует"),
    ("Класс 1", f"NI 0–{tau1} — низкая потребность"),
    ("Класс 2", f"NI {tau1}–{tau2} — средняя потребность"),
    ("Класс 3", f"NI > {tau2} — высокая потребность"),
    ("", ""),
    ("ПОКАЗАТЕЛИ ТАБЛИЦЫ", ""),
    ("NeedIndex (ср.)",   "Среднее NI по всем ячейкам квартала"),
    ("% активных ячеек", "Доля ячеек квартала с NI > 0"),
]
for ri, (lbl, val) in enumerate(legend_rows, 1):
    is_header = val == "" and lbl != ""
    c1 = wl.cell(row=ri, column=1, value=lbl)
    c2 = wl.cell(row=ri, column=2, value=val)
    if is_header:
        c1.font = fnt(bold=True, color=H_FG); c1.fill = hfill(H_BG)
        c2.fill = hfill(H_BG)
    else:
        c1.font = fnt(bold=True); c2.font = fnt()
    for c in (c1, c2):
        c.border = brd(); c.alignment = aln("left", wrap=True)
    wl.row_dimensions[ri].height = 16

print('  Лист "Легенда" — готов')

wb.save(OUTPUT_FILE)
print(f"\nГотово: {OUTPUT_FILE}")